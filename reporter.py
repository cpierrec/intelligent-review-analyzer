"""
reporter.py - Kurumsal Raporlama Katmanı
=========================================
Amaç: İşlenmiş ve analiz edilmiş verileri iş birimlerinin okuyabileceği
      formatlara (JSON, CSV, Excel/xlsx) dönüştürerek diske kaydetmek.

Mimari Kararlar:
  - Her çıktı formatı için ayrı metod (Single Responsibility Principle).
  - Dosya adlarına UTC zaman damgası eklenerek versiyon karışıklığı önlenir.
  - Excel exportu için openpyxl kullanılır; renk kodlaması ve sütun genişliği
    ile kurumsal görünüm sağlanır (sadece xlsxwriter yerine tam kontrol).
  - Tüm yazma işlemleri try-except ile korunur; bir format başarısız olursa
    diğerleri çalışmaya devam eder (partial success senaryosu).
  - Dosya boyutu ve satır sayısı loglanarak audit trail oluşturulur.
"""

from __future__ import annotations

import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from models import AnalyticsSummary
from config import ReportConfig

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Excel Stil Sabitleri (openpyxl renk kodları)
# ---------------------------------------------------------------------------

# Sentiment'e göre satır arka plan rengi
SENTIMENT_COLORS = {
    "positive": "C8E6C9",   # Açık yeşil
    "negative": "FFCDD2",   # Açık kırmızı
    "neutral":  "F5F5F5",   # Açık gri
    "mixed":    "FFF9C4",   # Açık sarı
}

# Urgency seviyesine göre hücre rengi
URGENCY_COLORS = {
    "critical": "B71C1C",   # Koyu kırmızı (metin beyaz)
    "high":     "FF8F00",   # Amber
    "medium":   "F9A825",   # Sarı
    "low":      "388E3C",   # Yeşil
}

# Header satırı arka plan rengi
HEADER_COLOR = "1565C0"     # Koyu mavi
HEADER_FONT_COLOR = "FFFFFF"  # Beyaz

# Sütun genişlik haritası
COLUMN_WIDTHS = {
    "source_id": 15,
    "product_name": 35,
    "product_category": 18,
    "sentiment": 12,
    "sentiment_score": 16,
    "confidence_score": 16,
    "reviewer_rating": 15,
    "urgency_level": 15,
    "action_required": 15,
    "summary": 60,
    "key_topics_str": 40,
    "pros_str": 40,
    "cons_str": 40,
    "language_detected": 15,
    "data_quality": 14,
    "analyzed_at": 22,
}


# ---------------------------------------------------------------------------
# Ana Reporter Sınıfı
# ---------------------------------------------------------------------------

class DataReporter:
    """
    Analiz edilmiş verileri kurumsal raporlara dönüştüren sınıf.

    Desteklenen çıktı formatları:
      - JSON  : Ham ve özet verileri yapılandırılmış JSON olarak
      - CSV   : Düz tablo formatı, her analitik araçla uyumlu
      - Excel : Renkli, formatlı kurumsal rapor (openpyxl)

    Kullanım:
        reporter = DataReporter(report_cfg)
        paths = reporter.export_all(clean_df, summary)
    """

    def __init__(self, config: ReportConfig) -> None:
        self._config = config
        self._output_dir = Path(config.output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(
            f"DataReporter başlatıldı: "
            f"çıktı dizini={self._output_dir.resolve()} | "
            f"formatlar={config.export_formats}"
        )

    def _build_filename(self, base_name: str, extension: str) -> Path:
        """
        Zaman damgalı dosya adı oluşturur.
        Örnek: reviews_analysis_20240615_143022.xlsx
        """
        if self._config.timestamp_filenames:
            ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"{base_name}_{ts}.{extension}"
        else:
            filename = f"{base_name}.{extension}"
        return self._output_dir / filename

    def _get_export_columns(self, df: pd.DataFrame) -> list[str]:
        """
        Raporda yer alacak sütunları öncelik sırasına göre döndürür.
        Varsa göster, yoksa atla (esnek sütun seçimi).
        """
        priority_columns = [
            "source_id",
            "product_name",
            "product_category",
            "sentiment",
            "sentiment_score",
            "confidence_score",
            "reviewer_rating",
            "urgency_level",
            "action_required",
            "action_notes",
            "summary",
            "key_topics_str",
            "pros_str",
            "cons_str",
            "is_verified_purchase",
            "reviewer_expertise",
            "mentioned_competitors",
            "language_detected",
            "data_quality",
            "is_spam_or_fake",
            "analyzed_at",
        ]
        return [col for col in priority_columns if col in df.columns]

    # -----------------------------------------------------------------------
    # JSON Export
    # -----------------------------------------------------------------------

    def export_json(
        self,
        clean_df: pd.DataFrame,
        summary: AnalyticsSummary,
    ) -> Optional[Path]:
        """
        İki bölümlü JSON raporu oluşturur:
          1. "summary"  : AnalyticsSummary KPI özeti
          2. "records"  : Tüm analiz edilmiş kayıtlar (liste)

        Neden iki bölüm? Tüketen servis sadece KPI'ya ihtiyaç duyabilir
        (summary), veya tüm kayıtları işlemek isteyebilir (records).
        """
        output_path = self._build_filename("reviews_analysis", "json")
        logger.info(f"JSON export başlıyor: {output_path}")

        try:
            # DataFrame'i JSON-serializable formata çevir
            records: list[dict[str, Any]] = []
            if not clean_df.empty:
                export_cols = self._get_export_columns(clean_df)
                records = (
                    clean_df[export_cols]
                    .fillna("")
                    .to_dict(orient="records")
                )
                # Boolean alanları düzelt (numpy bool → Python bool)
                for rec in records:
                    for k, v in rec.items():
                        if hasattr(v, "item"):   # numpy scalar
                            rec[k] = v.item()

            # Özet modeli dict'e çevir
            summary_dict = json.loads(
                summary.model_dump_json(indent=None)
            )

            output_data = {
                "meta": {
                    "engine": "AI-Powered Data Integration Engine v1.0",
                    "exported_at": datetime.now(timezone.utc).isoformat(),
                    "total_records": len(records),
                    "report_id": summary.report_id,
                },
                "summary": summary_dict,
                "records": records,
            }

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(
                    output_data,
                    f,
                    indent=self._config.json_indent,
                    ensure_ascii=False,
                    default=str,  # datetime ve diğer serialize edilemeyen tipler için
                )

            file_size_kb = output_path.stat().st_size / 1024
            logger.info(
                f"✓ JSON export tamamlandı: {output_path.name} "
                f"({file_size_kb:.1f} KB, {len(records)} kayıt)"
            )
            return output_path

        except Exception as e:
            logger.error(f"✗ JSON export başarısız: {e}", exc_info=True)
            return None

    # -----------------------------------------------------------------------
    # CSV Export
    # -----------------------------------------------------------------------

    def export_csv(self, clean_df: pd.DataFrame) -> Optional[Path]:
        """
        UTF-8 BOM ile CSV export (Excel'de Türkçe karakter desteği için).
        utf-8-sig encoding: Excel'in BOM'u otomatik tanımasını sağlar.
        """
        output_path = self._build_filename("reviews_analysis", "csv")
        logger.info(f"CSV export başlıyor: {output_path}")

        try:
            if clean_df.empty:
                logger.warning("Boş DataFrame, boş CSV oluşturuluyor.")
                pd.DataFrame().to_csv(output_path, index=False, encoding="utf-8-sig")
                return output_path

            export_cols = self._get_export_columns(clean_df)
            export_df = clean_df[export_cols].copy()

            # Sayısal sütunları iki ondalık basamağa yuvarla
            float_cols = export_df.select_dtypes(include=["float64", "float32"]).columns
            export_df[float_cols] = export_df[float_cols].round(4)

            export_df.to_csv(
                output_path,
                index=False,
                encoding="utf-8-sig",   # Excel BOM uyumluluğu
                float_format="%.4f",
            )

            file_size_kb = output_path.stat().st_size / 1024
            logger.info(
                f"✓ CSV export tamamlandı: {output_path.name} "
                f"({file_size_kb:.1f} KB, {len(export_df)} satır)"
            )
            return output_path

        except Exception as e:
            logger.error(f"✗ CSV export başarısız: {e}", exc_info=True)
            return None

    # -----------------------------------------------------------------------
    # Excel Export
    # -----------------------------------------------------------------------

    def export_excel(
        self,
        clean_df: pd.DataFrame,
        summary: AnalyticsSummary,
    ) -> Optional[Path]:
        """
        Kurumsal görünümlü, çok sayfalı Excel raporu oluşturur.

        Sayfalar:
          1. "Review Analysis"  : Renk kodlu, formatlı detay tablosu
          2. "KPI Summary"      : Yönetici özeti (key metrics)
          3. "Sentiment Chart"  : Sentiment dağılım verisi (grafikler için hazır)

        openpyxl doğrudan kullanılarak piksel-hassas formatlama yapılır.
        """
        output_path = self._build_filename("reviews_analysis", "xlsx")
        logger.info(f"Excel export başlıyor: {output_path}")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                GradientFill,
                PatternFill,
                Side,
            )
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.dataframe import dataframe_to_rows

        except ImportError:
            logger.warning(
                "openpyxl bulunamadı. Excel export atlandı. "
                "Kurulum: pip install openpyxl"
            )
            return None

        try:
            wb = Workbook()

            # ── Sayfa 1: Detay Tablosu ──────────────────────────────────────
            ws_detail = wb.active
            ws_detail.title = self._config.excel_sheet_name

            if not clean_df.empty:
                export_cols = self._get_export_columns(clean_df)
                export_df = clean_df[export_cols].copy()

                # Float yuvarla
                float_cols = export_df.select_dtypes(include=["float64", "float32"]).columns
                export_df[float_cols] = export_df[float_cols].round(4)

                # List/dict hücrelerini string'e çevir - openpyxl [] değerini kabul etmez
                for col in export_df.columns:
                    export_df[col] = export_df[col].apply(
                        lambda v: (
                            "|".join(str(x) for x in v) if isinstance(v, list)
                            else (str(v) if isinstance(v, dict) else v)
                        )
                    )

                # Sütun başlıklarını insan okunabilir formata çevir
                export_df.columns = [
                    col.replace("_", " ").replace("str", "").title().strip()
                    for col in export_df.columns
                ]

                # Header stili
                header_fill = PatternFill(
                    start_color=HEADER_COLOR,
                    end_color=HEADER_COLOR,
                    fill_type="solid",
                )
                header_font = Font(
                    name="Calibri",
                    bold=True,
                    color=HEADER_FONT_COLOR,
                    size=11,
                )
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # DataFrame satırlarını çalışma sayfasına yaz
                for row_idx, row in enumerate(
                    dataframe_to_rows(export_df, index=False, header=True), start=1
                ):
                    ws_detail.append(row)

                    if row_idx == 1:
                        # Header satırını formatla
                        for cell in ws_detail[row_idx]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center", wrap_text=True
                            )
                            cell.border = thin_border
                    else:
                        # Veri satırlarını formatla
                        # Sentiment sütununu bul ve renklendir
                        sentiment_col_idx = None
                        urgency_col_idx = None
                        for col_idx, col_name in enumerate(export_df.columns, start=1):
                            col_lower = col_name.lower()
                            if "sentiment" == col_lower.strip():
                                sentiment_col_idx = col_idx
                            elif "urgency" in col_lower:
                                urgency_col_idx = col_idx

                        for cell in ws_detail[row_idx]:
                            cell.border = thin_border
                            cell.alignment = Alignment(
                                vertical="top", wrap_text=True
                            )

                        # Sentiment renklendirme
                        if sentiment_col_idx:
                            sentiment_cell = ws_detail.cell(row_idx, sentiment_col_idx)
                            sentiment_val = str(sentiment_cell.value or "").lower()
                            color = SENTIMENT_COLORS.get(sentiment_val, "FFFFFF")
                            sentiment_cell.fill = PatternFill(
                                start_color=color,
                                end_color=color,
                                fill_type="solid",
                            )

                # Sütun genişliklerini ayarla
                for col_idx, col_name in enumerate(export_df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    # Orijinal sütun adından genişlik eşleştirmesi dene
                    orig_col = export_cols[col_idx - 1] if col_idx - 1 < len(export_cols) else ""
                    width = COLUMN_WIDTHS.get(orig_col, 20)
                    ws_detail.column_dimensions[col_letter].width = width

                # Satır yüksekliği
                ws_detail.row_dimensions[1].height = 35
                for row_num in range(2, len(export_df) + 2):
                    ws_detail.row_dimensions[row_num].height = 25

                # Freeze panes - üst satır ve ilk sütun sabit
                ws_detail.freeze_panes = "B2"

                # Auto filter
                ws_detail.auto_filter.ref = ws_detail.dimensions

            # ── Sayfa 2: KPI Özeti ─────────────────────────────────────────
            ws_kpi = wb.create_sheet("KPI Summary")
            self._write_kpi_sheet(ws_kpi, summary)

            # ── Sayfa 3: Sentiment Dağılım Verisi ─────────────────────────
            ws_chart = wb.create_sheet("Sentiment Distribution")
            self._write_sentiment_sheet(ws_chart, summary)

            wb.save(output_path)

            file_size_kb = output_path.stat().st_size / 1024
            logger.info(
                f"✓ Excel export tamamlandı: {output_path.name} "
                f"({file_size_kb:.1f} KB, {len(clean_df)} satır, 3 sayfa)"
            )
            return output_path

        except Exception as e:
            logger.error(f"✗ Excel export başarısız: {e}", exc_info=True)
            return None

    def _write_kpi_sheet(self, ws, summary: AnalyticsSummary) -> None:
        """KPI özet sayfasını yazar."""
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return

        title_font = Font(name="Calibri", bold=True, size=14, color="1565C0")
        label_font = Font(name="Calibri", bold=True, size=11)
        value_font = Font(name="Calibri", size=11)
        header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

        rows = [
            ("AI DATA ENGINE - KPI RAPORU", ""),
            (f"Rapor ID: {summary.report_id}", ""),
            (f"Oluşturulma: {summary.generated_at.strftime('%Y-%m-%d %H:%M UTC')}", ""),
            ("", ""),
            ("METRİK", "DEĞER"),
            ("Toplam Analiz Edilen Yorum", summary.total_reviews_analyzed),
            ("Yüksek Kaliteli Kayıt", summary.high_quality_reviews),
            ("Spam/Sahte Filtrelenen", summary.spam_excluded),
            ("Ortalama Sentiment Skoru", f"{summary.average_sentiment_score:+.4f}"),
            ("Ortalama Müşteri Puanı", f"{summary.average_rating:.2f}/5.0" if summary.average_rating else "N/A"),
            ("Ortalama AI Güven Skoru", f"{summary.average_confidence:.1%}"),
            ("Aksiyon Gerektiren Kayıt", summary.action_required_count),
            ("CRITICAL Öncelikli", summary.critical_urgency_count),
            ("HIGH Öncelikli", summary.high_urgency_count),
            ("", ""),
            ("DUYGU DAĞILIMI", ""),
        ]

        for sentiment, count in summary.sentiment_distribution.items():
            pct = count / max(summary.total_reviews_analyzed, 1)
            rows.append((f"  {sentiment.title()}", f"{count} ({pct:.1%})"))

        rows += [("", ""), ("KATEGORİ DAĞILIMI", "")]
        for cat, count in summary.category_distribution.items():
            rows.append((f"  {cat.replace('_', ' ').title()}", count))

        for row_idx, (label, value) in enumerate(rows, start=1):
            cell_a = ws.cell(row=row_idx, column=1, value=label)
            cell_b = ws.cell(row=row_idx, column=2, value=value)

            if row_idx == 1:
                cell_a.font = title_font
            elif label in ("METRİK", "DUYGU DAĞILIMI", "KATEGORİ DAĞILIMI"):
                cell_a.font = label_font
                cell_a.fill = header_fill
                cell_b.fill = header_fill
            else:
                cell_a.font = label_font
                cell_b.font = value_font

            cell_a.alignment = Alignment(vertical="center")
            cell_b.alignment = Alignment(vertical="center", horizontal="right")
            ws.row_dimensions[row_idx].height = 20

    def _write_sentiment_sheet(self, ws, summary: AnalyticsSummary) -> None:
        """Sentiment dağılım verisi sayfasını yazar (grafik için ham veri)."""
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            return

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

        # Başlık
        ws.cell(1, 1, "Sentiment").font = Font(bold=True)
        ws.cell(1, 2, "Count").font = Font(bold=True)
        ws.cell(1, 3, "Percentage").font = Font(bold=True)

        total = max(summary.total_reviews_analyzed, 1)
        for row_idx, (sentiment, count) in enumerate(
            summary.sentiment_distribution.items(), start=2
        ):
            ws.cell(row_idx, 1, sentiment.title())
            ws.cell(row_idx, 2, count)
            ws.cell(row_idx, 3, round(count / total, 4))

        # Bar chart ekle
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Sentiment Distribution"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Sentiment"
            chart.style = 10

            data = Reference(
                ws,
                min_col=2,
                min_row=1,
                max_row=len(summary.sentiment_distribution) + 1,
            )
            cats = Reference(
                ws,
                min_col=1,
                min_row=2,
                max_row=len(summary.sentiment_distribution) + 1,
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws.add_chart(chart, "E2")
        except Exception as chart_err:
            logger.debug(f"Grafik eklenemedi (non-critical): {chart_err}")

    # -----------------------------------------------------------------------
    # Ana Export Metodu
    # -----------------------------------------------------------------------

    def export_all(
        self,
        clean_df: pd.DataFrame,
        summary: AnalyticsSummary,
    ) -> dict[str, Optional[Path]]:
        """
        Tüm yapılandırılmış formatları sırayla export eder.

        Bir format başarısız olsa bile diğerleri çalışmaya devam eder.
        Döndürülen dict: {"json": Path|None, "csv": Path|None, "xlsx": Path|None}
        """
        logger.info(
            f"━━━ Export Başlıyor ━━━ "
            f"formatlar={self._config.export_formats} | "
            f"kayıt={len(clean_df)}"
        )

        results: dict[str, Optional[Path]] = {}

        if "json" in self._config.export_formats:
            results["json"] = self.export_json(clean_df, summary)

        if "csv" in self._config.export_formats:
            results["csv"] = self.export_csv(clean_df)

        if "xlsx" in self._config.export_formats:
            results["xlsx"] = self.export_excel(clean_df, summary)

        # Özet log
        successful_exports = [fmt for fmt, path in results.items() if path]
        failed_exports = [fmt for fmt, path in results.items() if not path]

        logger.info(
            f"━━━ Export Tamamlandı ━━━ "
            f"başarılı={successful_exports} | "
            f"başarısız={failed_exports}"
        )

        if successful_exports:
            logger.info("📁 Üretilen dosyalar:")
            for fmt, path in results.items():
                if path:
                    size_kb = path.stat().st_size / 1024
                    logger.info(f"   {fmt.upper():5} → {path}  ({size_kb:.1f} KB)")

        return results
